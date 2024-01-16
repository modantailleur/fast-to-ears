#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Oct 31 14:12:06 2022

@author: user
"""

import torch
import librosa
import numpy as np
import matplotlib.pyplot as plt
import openpyxl
from pathlib import Path
from sklearn import preprocessing

from yamnet.torch_audioset.yamnet.model import yamnet as torch_yamnet
from yamnet.torch_audioset.yamnet.model import yamnet_category_metadata

import torch.nn.functional as F
from pathlib import Path
import utils.util as ut

class YamnetMelInference():
    def __init__(self, device=torch.device("cpu"), verbose=False):
        self.name = "YamNet"
        self.sample_rate = 32000
        self.window_size = 1024
        self.hop_size = 320
        self.mel_bins = 64
        self.fmin = 50
        self.fmax = 14000
        #self.labels_str = ['Speech', 'Male speech, man speaking', 'Female speech, woman speaking', 'Child speech, kid speaking', 'Conversation', 'Narration, monologue', 'Babbling', 'Speech synthesizer', 'Shout', 'Bellow', 'Whoop', 'Yell', 'Battle cry', 'Children shouting', 'Screaming', 'Whispering', 'Laughter', 'Baby laughter', 'Giggle', 'Snicker', 'Belly laugh', 'Chuckle, chortle', 'Crying, sobbing', 'Baby cry, infant cry', 'Whimper', 'Wail, moan', 'Sigh', 'Singing', 'Choir', 'Yodeling', 'Chant', 'Mantra', 'Male singing', 'Female singing', 'Child singing', 'Synthetic singing', 'Rapping', 'Humming', 'Groan', 'Grunt', 'Whistling', 'Breathing', 'Wheeze', 'Snoring', 'Gasp', 'Pant', 'Snort', 'Cough', 'Throat clearing', 'Sneeze', 'Sniff', 'Run', 'Shuffle', 'Walk, footsteps', 'Chewing, mastication', 'Biting', 'Gargling', 'Stomach rumble', 'Burping, eructation', 'Hiccup', 'Fart', 'Hands', 'Finger snapping', 'Clapping', 'Heart sounds, heartbeat', 'Heart murmur', 'Cheering', 'Applause', 'Chatter', 'Crowd', 'Hubbub, speech noise, speech babble', 'Children playing', 'Animal', 'Domestic animals, pets', 'Dog', 'Bark', 'Yip', 'Howl', 'Bow-wow', 'Growling', 'Whimper (dog)', 'Cat', 'Purr', 'Meow', 'Hiss', 'Caterwaul', 'Livestock, farm animals, working animals', 'Horse', 'Clip-clop', 'Neigh, whinny', 'Cattle, bovinae', 'Moo', 'Cowbell', 'Pig', 'Oink', 'Goat', 'Bleat', 'Sheep', 'Fowl', 'Chicken, rooster', 'Cluck', 'Crowing, cock-a-doodle-doo', 'Turkey', 'Gobble', 'Duck', 'Quack', 'Goose', 'Honk', 'Wild animals', 'Roaring cats (lions, tigers)', 'Roar', 'Bird', 'Bird vocalization, bird call, bird song', 'Chirp, tweet', 'Squawk', 'Pigeon, dove', 'Coo', 'Crow', 'Caw', 'Owl', 'Hoot', 'Bird flight, flapping wings', 'Canidae, dogs, wolves', 'Rodents, rats, mice', 'Mouse', 'Patter', 'Insect', 'Cricket', 'Mosquito', 'Fly, housefly', 'Buzz', 'Bee, wasp, etc.', 'Frog', 'Croak', 'Snake', 'Rattle', 'Whale vocalization', 'Music', 'Musical instrument', 'Plucked string instrument', 'Guitar', 'Electric guitar', 'Bass guitar', 'Acoustic guitar', 'Steel guitar, slide guitar', 'Tapping (guitar technique)', 'Strum', 'Banjo', 'Sitar', 'Mandolin', 'Zither', 'Ukulele', 'Keyboard (musical)', 'Piano', 'Electric piano', 'Organ', 'Electronic organ', 'Hammond organ', 'Synthesizer', 'Sampler', 'Harpsichord', 'Percussion', 'Drum kit', 'Drum machine', 'Drum', 'Snare drum', 'Rimshot', 'Drum roll', 'Bass drum', 'Timpani', 'Tabla', 'Cymbal', 'Hi-hat', 'Wood block', 'Tambourine', 'Rattle (instrument)', 'Maraca', 'Gong', 'Tubular bells', 'Mallet percussion', 'Marimba, xylophone', 'Glockenspiel', 'Vibraphone', 'Steelpan', 'Orchestra', 'Brass instrument', 'French horn', 'Trumpet', 'Trombone', 'Bowed string instrument', 'String section', 'Violin, fiddle', 'Pizzicato', 'Cello', 'Double bass', 'Wind instrument, woodwind instrument', 'Flute', 'Saxophone', 'Clarinet', 'Harp', 'Bell', 'Church bell', 'Jingle bell', 'Bicycle bell', 'Tuning fork', 'Chime', 'Wind chime', 'Change ringing (campanology)', 'Harmonica', 'Accordion', 'Bagpipes', 'Didgeridoo', 'Shofar', 'Theremin', 'Singing bowl', 'Scratching (performance technique)', 'Pop music', 'Hip hop music', 'Beatboxing', 'Rock music', 'Heavy metal', 'Punk rock', 'Grunge', 'Progressive rock', 'Rock and roll', 'Psychedelic rock', 'Rhythm and blues', 'Soul music', 'Reggae', 'Country', 'Swing music', 'Bluegrass', 'Funk', 'Folk music', 'Middle Eastern music', 'Jazz', 'Disco', 'Classical music', 'Opera', 'Electronic music', 'House music', 'Techno', 'Dubstep', 'Drum and bass', 'Electronica', 'Electronic dance music', 'Ambient music', 'Trance music', 'Music of Latin America', 'Salsa music', 'Flamenco', 'Blues', 'Music for children', 'New-age music', 'Vocal music', 'A capella', 'Music of Africa', 'Afrobeat', 'Christian music', 'Gospel music', 'Music of Asia', 'Carnatic music', 'Music of Bollywood', 'Ska', 'Traditional music', 'Independent music', 'Song', 'Background music', 'Theme music', 'Jingle (music)', 'Soundtrack music', 'Lullaby', 'Video game music', 'Christmas music', 'Dance music', 'Wedding music', 'Happy music', 'Funny music', 'Sad music', 'Tender music', 'Exciting music', 'Angry music', 'Scary music', 'Wind', 'Rustling leaves', 'Wind noise (microphone)', 'Thunderstorm', 'Thunder', 'Water', 'Rain', 'Raindrop', 'Rain on surface', 'Stream', 'Waterfall', 'Ocean', 'Waves, surf', 'Steam', 'Gurgling', 'Fire', 'Crackle', 'Vehicle', 'Boat, Water vehicle', 'Sailboat, sailing ship', 'Rowboat, canoe, kayak', 'Motorboat, speedboat', 'Ship', 'Motor vehicle (road)', 'Car', 'Vehicle horn, car horn, honking', 'Toot', 'Car alarm', 'Power windows, electric windows', 'Skidding', 'Tire squeal', 'Car passing by', 'Race car, auto racing', 'Truck', 'Air brake', 'Air horn, truck horn', 'Reversing beeps', 'Ice cream truck, ice cream van', 'Bus', 'Emergency vehicle', 'Police car (siren)', 'Ambulance (siren)', 'Fire engine, fire truck (siren)', 'Motorcycle', 'Traffic noise, roadway noise', 'Rail transport', 'Train', 'Train whistle', 'Train horn', 'Railroad car, train wagon', 'Train wheels squealing', 'Subway, metro, underground', 'Aircraft', 'Aircraft engine', 'Jet engine', 'Propeller, airscrew', 'Helicopter', 'Fixed-wing aircraft, airplane', 'Bicycle', 'Skateboard', 'Engine', 'Light engine (high frequency)', "Dental drill, dentist's drill", 'Lawn mower', 'Chainsaw', 'Medium engine (mid frequency)', 'Heavy engine (low frequency)', 'Engine knocking', 'Engine starting', 'Idling', 'Accelerating, revving, vroom', 'Door', 'Doorbell', 'Ding-dong', 'Sliding door', 'Slam', 'Knock', 'Tap', 'Squeak', 'Cupboard open or close', 'Drawer open or close', 'Dishes, pots, and pans', 'Cutlery, silverware', 'Chopping (food)', 'Frying (food)', 'Microwave oven', 'Blender', 'Water tap, faucet', 'Sink (filling or washing)', 'Bathtub (filling or washing)', 'Hair dryer', 'Toilet flush', 'Toothbrush', 'Electric toothbrush', 'Vacuum cleaner', 'Zipper (clothing)', 'Keys jangling', 'Coin (dropping)', 'Scissors', 'Electric shaver, electric razor', 'Shuffling cards', 'Typing', 'Typewriter', 'Computer keyboard', 'Writing', 'Alarm', 'Telephone', 'Telephone bell ringing', 'Ringtone', 'Telephone dialing, DTMF', 'Dial tone', 'Busy signal', 'Alarm clock', 'Siren', 'Civil defense siren', 'Buzzer', 'Smoke detector, smoke alarm', 'Fire alarm', 'Foghorn', 'Whistle', 'Steam whistle', 'Mechanisms', 'Ratchet, pawl', 'Clock', 'Tick', 'Tick-tock', 'Gears', 'Pulleys', 'Sewing machine', 'Mechanical fan', 'Air conditioning', 'Cash register', 'Printer', 'Camera', 'Single-lens reflex camera', 'Tools', 'Hammer', 'Jackhammer', 'Sawing', 'Filing (rasp)', 'Sanding', 'Power tool', 'Drill', 'Explosion', 'Gunshot, gunfire', 'Machine gun', 'Fusillade', 'Artillery fire', 'Cap gun', 'Fireworks', 'Firecracker', 'Burst, pop', 'Eruption', 'Boom', 'Wood', 'Chop', 'Splinter', 'Crack', 'Glass', 'Chink, clink', 'Shatter', 'Liquid', 'Splash, splatter', 'Slosh', 'Squish', 'Drip', 'Pour', 'Trickle, dribble', 'Gush', 'Fill (with liquid)', 'Spray', 'Pump (liquid)', 'Stir', 'Boiling', 'Sonar', 'Arrow', 'Whoosh, swoosh, swish', 'Thump, thud', 'Thunk', 'Electronic tuner', 'Effects unit', 'Chorus effect', 'Basketball bounce', 'Bang', 'Slap, smack', 'Whack, thwack', 'Smash, crash', 'Breaking', 'Bouncing', 'Whip', 'Flap', 'Scratch', 'Scrape', 'Rub', 'Roll', 'Crushing', 'Crumpling, crinkling', 'Tearing', 'Beep, bleep', 'Ping', 'Ding', 'Clang', 'Squeal', 'Creak', 'Rustle', 'Whir', 'Clatter', 'Sizzle', 'Clicking', 'Clickety-clack', 'Rumble', 'Plop', 'Jingle, tinkle', 'Hum', 'Zing', 'Boing', 'Crunch', 'Silence', 'Sine wave', 'Harmonic', 'Chirp tone', 'Sound effect', 'Pulse', 'Inside, small room', 'Inside, large room or hall', 'Inside, public space', 'Outside, urban or manmade', 'Outside, rural or natural', 'Reverberation', 'Echo', 'Noise', 'Environmental noise', 'Static', 'Mains hum', 'Distortion', 'Sidetone', 'Cacophony', 'White noise', 'Pink noise', 'Throbbing', 'Vibration', 'Television', 'Radio', 'Field recording']
        self.labels_str = [x['name'] for x in yamnet_category_metadata()]

        self.le = preprocessing.LabelEncoder()
        self.labels_enc = self.le.fit_transform(self.labels_str)
        self.labels_enc = torch.from_numpy(self.labels_enc)

        self.n_labels = len(self.labels_enc)
        self.device = device
        
        ###############
        #models loading
        self.model = torch_yamnet(pretrained=False)
        # Manually download the `yamnet.pth` file.
        self.model.load_state_dict(torch.load(Path().absolute() / 'yamnet' / 'yamnet.pth', map_location=device))
        self.model.to(device)

        if verbose:
            print('YamNet Parameters')
            ut.count_parameters(self.model)
            
        ###############
        #sub-labels
        sub_classes_path = './utils/sub_classes.xlsx'
        self.sub_classes_dict = open_subclasses_dict(sub_classes_path)
        self.labels_tvb_str = [label for label in self.labels_str if self.sub_classes_dict[label] in ['t', 'v', 'b'] ]
        self.labels_tvb_enc = self.le.transform(self.labels_tvb_str)
        self.labels_tvb_enc = torch.from_numpy(self.labels_tvb_enc)
        
        self.n_labels_tvb = len(self.labels_tvb_enc)

        ###############
        #thresholds (temporary, waiting for mail answer)
        self.threshold = [0.2] * self.n_labels

    # def inference(self, x, filter_classes=True):
    #     # Forward
    #     with torch.no_grad():
    #         self.mels_model.eval()
    #         batch_output_dict = self.mels_model(x, None)
        
    #     framewise_output = batch_output_dict['framewise_output'].data.cpu()[0]
    #     """(time_steps, classes_num)"""

    #     #print('Sound event detection result (time_steps x classes_num): {}'.format(
    #     #    framewise_output.shape))
        
    #     labels = self.labels
    #     if filter_classes == True:
    #         labels = self.labels_kept
    #         labels_kept_indices = [k for k in range(len(self.labels)) if self.sub_classes_dict[self.labels[k]] in ['t', 'v', 'b'] ]
    #         framewise_output = framewise_output[:, labels_kept_indices]
        
    #     print('AAAAAAAA')
    #     print(torch.max(framewise_output, axis=0)[0])
    #     sorted_indexes = torch.flip(torch.argsort(torch.max(framewise_output, axis=0)[0]), dims=[0])
    #     print(sorted_indexes)
    #     top_k = 10  # Show top results
    #     top_result_labels = np.array(labels)[sorted_indexes[0 : top_k]]
    #     top_result_mat = framewise_output[:, sorted_indexes[0 : top_k]]    
    #     """(time_steps, top_k)"""
        
    #     return(framewise_output, top_result_mat, top_result_labels)

    def simple_inference(self, x, filter_classes=True, softmax=False, no_grad=True, mean=True):
        if no_grad:
            with torch.no_grad():
                self.model.eval()
                # x = torch.from_numpy(patches)
                # x = x.unsqueeze(1)  # [5, 96, 64] -> [5, 1, 96, 64]
                #dimensions example: (2,1,96,64)
                logits = self.model(x, to_prob=True)
                #pt_pred = pt_pred.numpy()
        else:
            self.model.eval()
            logits = self.model(x, to_prob=True)
        return (logits)
        
    def inference(self, x, filter_classes=True, softmax=False, no_grad=True):
        # Forward
        
        if no_grad:
            with torch.no_grad():
                self.model.eval()
                # x = torch.from_numpy(patches)
                # x = x.unsqueeze(1)  # [5, 96, 64] -> [5, 1, 96, 64]
                #dimensions example: (2,1,96,64)
                logits = self.model(x, to_prob=True)
                #pt_pred = pt_pred.numpy()
        else:
            self.model.eval()
            logits = self.model(x, to_prob=True)
        
        logits = logits
        
        #print('Sound event detection result (time_steps x classes_num): {}'.format(
        #    framewise_output.shape))
        
        logits_tvb = torch.Tensor([])
        
        labels_enc = self.labels_enc
        if filter_classes == True:
            labels_enc = self.labels_tvb_enc
            labels_tvb_enc_indices = [k for k in range(len(self.labels_enc)) if self.sub_classes_dict[self.labels_str[k]] in ['t', 'v', 'b'] ]
            logits_tvb = logits[:, labels_tvb_enc_indices]

        if softmax:
            logits = F.log_softmax(logits, dim=1)
            logits_tvb = F.log_softmax(logits_tvb, dim=1)
            #scores = F.softmax(scores, dim=1)
        
        
        #use for showing top_results

        # labels = labels.to(self.device)
        
        # #sorted_indexes = torch.flip(torch.argsort(torch.max(scores, dim=0)[0]), dims=[0])
        # sorted_indexes = torch.flip(torch.argsort(scores), dims=[0,1])

        # top_k = 10  # Show top results
        
        # top_result_labels = labels[sorted_indexes[:, 0 : top_k]]
        # top_result_mat = scores[:, sorted_indexes[:, 0 : top_k]]    

        # """(time_steps, top_k)"""
        
        # #print(self.le.inverse_transform(top_result_labels[0]))
               
        return(logits, logits_tvb) 
    
    # def inference(self, x, filter_classes=True, device=torch.device("cpu")):
    #     # Forward
    #     with torch.no_grad():
    #         self.mels_model.eval()
    #         batch_output_dict = self.mels_model(x, None)
        
    #     framewise_output = batch_output_dict['framewise_output'].data.cpu().numpy()[0]
    #     """(time_steps, classes_num)"""

    #     #print('Sound event detection result (time_steps x classes_num): {}'.format(
    #     #    framewise_output.shape))
        
    #     labels = self.labels
    #     if filter_classes == True:
    #         labels = self.labels_kept
    #         labels_kept_indices = [k for k in range(len(self.labels)) if self.sub_classes_dict[self.labels[k]] in ['t', 'v', 'b'] ]
    #         framewise_output = framewise_output[:, labels_kept_indices]
    #     sorted_indexes = np.argsort(np.max(framewise_output, axis=0))[::-1]
    #     top_k = 10  # Show top results
    #     top_result_labels = np.array(labels)[sorted_indexes[0 : top_k]]
    #     top_result_mat = framewise_output[:, sorted_indexes[0 : top_k]]    
    #     """(time_steps, top_k)"""
        
    #     return(framewise_output, top_result_mat, top_result_labels)
    def logit_to_labels(self, input, tvb=False):
        
        #average over a whole file
        logits_tvb = input.mean(dim=1)
        
        #use for showing top_results
        if tvb:
            labels_enc = self.labels_tvb_enc
        else:
            labels_enc = self.labels_enc
            
        #labels_enc = labels_enc.to(self.device)

        sorted_indexes = torch.flip(torch.argsort(logits_tvb), dims=[1])

        top_k = 1  # Show top results
 
        labels_enc_top = labels_enc[sorted_indexes[:, 0 : top_k]]
        top_result_mat = logits_tvb[:, sorted_indexes[:, 0 : top_k]]    
        # """(time_steps, top_k)"""

        labels_enc_top = labels_enc_top

        labels_enc_top = labels_enc_top.flatten()
        labels_str_top = self.le.inverse_transform(labels_enc_top)

        return(labels_str_top)
    
    def logit_to_logit_tvb(self, logits):
        labels_enc = self.labels_tvb_enc
        labels_tvb_enc_indices = [k for k in range(len(self.labels_enc)) if self.sub_classes_dict[self.labels_str[k]] in ['t', 'v', 'b'] ]
        logits_tvb = logits[:,:, labels_tvb_enc_indices]
        return(logits_tvb)
    
    # def inference_from_scratch(self, file_name):
    #     wav_data, sr = librosa.load(file_name, sr=16000)
    #     waveform = wav_data
    #     #added by MT
    #     waveform = librosa.util.normalize(waveform)

    #     # Convert to mono and the sample rate expected by YAMNet.
    #     if len(waveform.shape) > 1:
    #         waveform = np.mean(waveform, axis=1)
        
    #     logits = self.simple_inference(waveform)

    #     return(logits)


def open_subclasses_dict(workbook_path):
    workbook = openpyxl.load_workbook(workbook_path)
    worksheet = workbook.active
    
    first_row = [] # The row where we stock the name of the column
    for col in range(1,2):
        first_row.append( worksheet.cell(1,col).value )
    # tronsform the workbook to a list of dictionnary
    sub_classes_dict = {}
    for row in range(2, worksheet.max_row+1):
        sub_classes_dict[worksheet.cell(row,1).value] = worksheet.cell(row,2).value
    return(sub_classes_dict)

# def pann_inference_mels(x_mels):
#     FORCE_CPU = False
#     useCuda = torch.cuda.is_available() and not FORCE_CPU
    
#     # Arugments & parameters
#     sample_rate = 32000
#     window_size = 1024
#     hop_size = 320
#     mel_bins = 64
#     fmin = 50
#     fmax = 14000
#     #model_type = 'Cnn14_DecisionLevelMax_mels'
#     checkpoint_path = 'pann/Cnn14_DecisionLevelMax_mAP=0.385.pth'
#     classes_num = 527
#     labels = ['Speech', 'Male speech, man speaking', 'Female speech, woman speaking', 'Child speech, kid speaking', 'Conversation', 'Narration, monologue', 'Babbling', 'Speech synthesizer', 'Shout', 'Bellow', 'Whoop', 'Yell', 'Battle cry', 'Children shouting', 'Screaming', 'Whispering', 'Laughter', 'Baby laughter', 'Giggle', 'Snicker', 'Belly laugh', 'Chuckle, chortle', 'Crying, sobbing', 'Baby cry, infant cry', 'Whimper', 'Wail, moan', 'Sigh', 'Singing', 'Choir', 'Yodeling', 'Chant', 'Mantra', 'Male singing', 'Female singing', 'Child singing', 'Synthetic singing', 'Rapping', 'Humming', 'Groan', 'Grunt', 'Whistling', 'Breathing', 'Wheeze', 'Snoring', 'Gasp', 'Pant', 'Snort', 'Cough', 'Throat clearing', 'Sneeze', 'Sniff', 'Run', 'Shuffle', 'Walk, footsteps', 'Chewing, mastication', 'Biting', 'Gargling', 'Stomach rumble', 'Burping, eructation', 'Hiccup', 'Fart', 'Hands', 'Finger snapping', 'Clapping', 'Heart sounds, heartbeat', 'Heart murmur', 'Cheering', 'Applause', 'Chatter', 'Crowd', 'Hubbub, speech noise, speech babble', 'Children playing', 'Animal', 'Domestic animals, pets', 'Dog', 'Bark', 'Yip', 'Howl', 'Bow-wow', 'Growling', 'Whimper (dog)', 'Cat', 'Purr', 'Meow', 'Hiss', 'Caterwaul', 'Livestock, farm animals, working animals', 'Horse', 'Clip-clop', 'Neigh, whinny', 'Cattle, bovinae', 'Moo', 'Cowbell', 'Pig', 'Oink', 'Goat', 'Bleat', 'Sheep', 'Fowl', 'Chicken, rooster', 'Cluck', 'Crowing, cock-a-doodle-doo', 'Turkey', 'Gobble', 'Duck', 'Quack', 'Goose', 'Honk', 'Wild animals', 'Roaring cats (lions, tigers)', 'Roar', 'Bird', 'Bird vocalization, bird call, bird song', 'Chirp, tweet', 'Squawk', 'Pigeon, dove', 'Coo', 'Crow', 'Caw', 'Owl', 'Hoot', 'Bird flight, flapping wings', 'Canidae, dogs, wolves', 'Rodents, rats, mice', 'Mouse', 'Patter', 'Insect', 'Cricket', 'Mosquito', 'Fly, housefly', 'Buzz', 'Bee, wasp, etc.', 'Frog', 'Croak', 'Snake', 'Rattle', 'Whale vocalization', 'Music', 'Musical instrument', 'Plucked string instrument', 'Guitar', 'Electric guitar', 'Bass guitar', 'Acoustic guitar', 'Steel guitar, slide guitar', 'Tapping (guitar technique)', 'Strum', 'Banjo', 'Sitar', 'Mandolin', 'Zither', 'Ukulele', 'Keyboard (musical)', 'Piano', 'Electric piano', 'Organ', 'Electronic organ', 'Hammond organ', 'Synthesizer', 'Sampler', 'Harpsichord', 'Percussion', 'Drum kit', 'Drum machine', 'Drum', 'Snare drum', 'Rimshot', 'Drum roll', 'Bass drum', 'Timpani', 'Tabla', 'Cymbal', 'Hi-hat', 'Wood block', 'Tambourine', 'Rattle (instrument)', 'Maraca', 'Gong', 'Tubular bells', 'Mallet percussion', 'Marimba, xylophone', 'Glockenspiel', 'Vibraphone', 'Steelpan', 'Orchestra', 'Brass instrument', 'French horn', 'Trumpet', 'Trombone', 'Bowed string instrument', 'String section', 'Violin, fiddle', 'Pizzicato', 'Cello', 'Double bass', 'Wind instrument, woodwind instrument', 'Flute', 'Saxophone', 'Clarinet', 'Harp', 'Bell', 'Church bell', 'Jingle bell', 'Bicycle bell', 'Tuning fork', 'Chime', 'Wind chime', 'Change ringing (campanology)', 'Harmonica', 'Accordion', 'Bagpipes', 'Didgeridoo', 'Shofar', 'Theremin', 'Singing bowl', 'Scratching (performance technique)', 'Pop music', 'Hip hop music', 'Beatboxing', 'Rock music', 'Heavy metal', 'Punk rock', 'Grunge', 'Progressive rock', 'Rock and roll', 'Psychedelic rock', 'Rhythm and blues', 'Soul music', 'Reggae', 'Country', 'Swing music', 'Bluegrass', 'Funk', 'Folk music', 'Middle Eastern music', 'Jazz', 'Disco', 'Classical music', 'Opera', 'Electronic music', 'House music', 'Techno', 'Dubstep', 'Drum and bass', 'Electronica', 'Electronic dance music', 'Ambient music', 'Trance music', 'Music of Latin America', 'Salsa music', 'Flamenco', 'Blues', 'Music for children', 'New-age music', 'Vocal music', 'A capella', 'Music of Africa', 'Afrobeat', 'Christian music', 'Gospel music', 'Music of Asia', 'Carnatic music', 'Music of Bollywood', 'Ska', 'Traditional music', 'Independent music', 'Song', 'Background music', 'Theme music', 'Jingle (music)', 'Soundtrack music', 'Lullaby', 'Video game music', 'Christmas music', 'Dance music', 'Wedding music', 'Happy music', 'Funny music', 'Sad music', 'Tender music', 'Exciting music', 'Angry music', 'Scary music', 'Wind', 'Rustling leaves', 'Wind noise (microphone)', 'Thunderstorm', 'Thunder', 'Water', 'Rain', 'Raindrop', 'Rain on surface', 'Stream', 'Waterfall', 'Ocean', 'Waves, surf', 'Steam', 'Gurgling', 'Fire', 'Crackle', 'Vehicle', 'Boat, Water vehicle', 'Sailboat, sailing ship', 'Rowboat, canoe, kayak', 'Motorboat, speedboat', 'Ship', 'Motor vehicle (road)', 'Car', 'Vehicle horn, car horn, honking', 'Toot', 'Car alarm', 'Power windows, electric windows', 'Skidding', 'Tire squeal', 'Car passing by', 'Race car, auto racing', 'Truck', 'Air brake', 'Air horn, truck horn', 'Reversing beeps', 'Ice cream truck, ice cream van', 'Bus', 'Emergency vehicle', 'Police car (siren)', 'Ambulance (siren)', 'Fire engine, fire truck (siren)', 'Motorcycle', 'Traffic noise, roadway noise', 'Rail transport', 'Train', 'Train whistle', 'Train horn', 'Railroad car, train wagon', 'Train wheels squealing', 'Subway, metro, underground', 'Aircraft', 'Aircraft engine', 'Jet engine', 'Propeller, airscrew', 'Helicopter', 'Fixed-wing aircraft, airplane', 'Bicycle', 'Skateboard', 'Engine', 'Light engine (high frequency)', "Dental drill, dentist's drill", 'Lawn mower', 'Chainsaw', 'Medium engine (mid frequency)', 'Heavy engine (low frequency)', 'Engine knocking', 'Engine starting', 'Idling', 'Accelerating, revving, vroom', 'Door', 'Doorbell', 'Ding-dong', 'Sliding door', 'Slam', 'Knock', 'Tap', 'Squeak', 'Cupboard open or close', 'Drawer open or close', 'Dishes, pots, and pans', 'Cutlery, silverware', 'Chopping (food)', 'Frying (food)', 'Microwave oven', 'Blender', 'Water tap, faucet', 'Sink (filling or washing)', 'Bathtub (filling or washing)', 'Hair dryer', 'Toilet flush', 'Toothbrush', 'Electric toothbrush', 'Vacuum cleaner', 'Zipper (clothing)', 'Keys jangling', 'Coin (dropping)', 'Scissors', 'Electric shaver, electric razor', 'Shuffling cards', 'Typing', 'Typewriter', 'Computer keyboard', 'Writing', 'Alarm', 'Telephone', 'Telephone bell ringing', 'Ringtone', 'Telephone dialing, DTMF', 'Dial tone', 'Busy signal', 'Alarm clock', 'Siren', 'Civil defense siren', 'Buzzer', 'Smoke detector, smoke alarm', 'Fire alarm', 'Foghorn', 'Whistle', 'Steam whistle', 'Mechanisms', 'Ratchet, pawl', 'Clock', 'Tick', 'Tick-tock', 'Gears', 'Pulleys', 'Sewing machine', 'Mechanical fan', 'Air conditioning', 'Cash register', 'Printer', 'Camera', 'Single-lens reflex camera', 'Tools', 'Hammer', 'Jackhammer', 'Sawing', 'Filing (rasp)', 'Sanding', 'Power tool', 'Drill', 'Explosion', 'Gunshot, gunfire', 'Machine gun', 'Fusillade', 'Artillery fire', 'Cap gun', 'Fireworks', 'Firecracker', 'Burst, pop', 'Eruption', 'Boom', 'Wood', 'Chop', 'Splinter', 'Crack', 'Glass', 'Chink, clink', 'Shatter', 'Liquid', 'Splash, splatter', 'Slosh', 'Squish', 'Drip', 'Pour', 'Trickle, dribble', 'Gush', 'Fill (with liquid)', 'Spray', 'Pump (liquid)', 'Stir', 'Boiling', 'Sonar', 'Arrow', 'Whoosh, swoosh, swish', 'Thump, thud', 'Thunk', 'Electronic tuner', 'Effects unit', 'Chorus effect', 'Basketball bounce', 'Bang', 'Slap, smack', 'Whack, thwack', 'Smash, crash', 'Breaking', 'Bouncing', 'Whip', 'Flap', 'Scratch', 'Scrape', 'Rub', 'Roll', 'Crushing', 'Crumpling, crinkling', 'Tearing', 'Beep, bleep', 'Ping', 'Ding', 'Clang', 'Squeal', 'Creak', 'Rustle', 'Whir', 'Clatter', 'Sizzle', 'Clicking', 'Clickety-clack', 'Rumble', 'Plop', 'Jingle, tinkle', 'Hum', 'Zing', 'Boing', 'Crunch', 'Silence', 'Sine wave', 'Harmonic', 'Chirp tone', 'Sound effect', 'Pulse', 'Inside, small room', 'Inside, large room or hall', 'Inside, public space', 'Outside, urban or manmade', 'Outside, rural or natural', 'Reverberation', 'Echo', 'Noise', 'Environmental noise', 'Static', 'Mains hum', 'Distortion', 'Sidetone', 'Cacophony', 'White noise', 'Pink noise', 'Throbbing', 'Vibration', 'Television', 'Radio', 'Field recording']
    
#     frames_per_second = sample_rate // hop_size
    
#     if useCuda:
#         print('Using CUDA.')
#         dtype = torch.cuda.FloatTensor
#         ltype = torch.cuda.LongTensor
#         #MT: add
#         device = torch.device("cuda:1")
#         print("cuda:1")
#     else:
#         print('No CUDA available.')
#         dtype = torch.FloatTensor
#         ltype = torch.LongTensor
#         #MT: add
#         device = torch.device("cpu")
        
        
#     # #open audio
#     # audio_path = 'traffic_short.wav'
#     # # Load audio
#     # (waveform, _) = librosa.core.load(audio_path, sr=sample_rate, mono=True)
    
#     # waveform = waveform[None, :]    # (1, audio_length)
#     # waveform = move_data_to_device(waveform, device)
    
    
#     # Model
    
#     #Model = eval(model_type)
#     mels_model = Cnn14_DecisionLevelMaxMels(sample_rate=sample_rate, window_size=window_size, 
#         hop_size=hop_size, mel_bins=mel_bins, fmin=fmin, fmax=fmax, 
#         classes_num=classes_num)
    
#     full_model =  Cnn14_DecisionLevelMax(sample_rate=sample_rate, window_size=window_size, 
#         hop_size=hop_size, mel_bins=mel_bins, fmin=fmin, fmax=fmax, 
#         classes_num=classes_num)
    

    
#     checkpoint = torch.load(checkpoint_path, map_location=device)
#     full_model.load_state_dict(checkpoint['model'])
    
#     full_dict = full_model.state_dict()
#     model_dict = mels_model.state_dict()
    
#     # 1. filter out unnecessary keys
#     full_dict = {k: v for k, v in full_dict.items() if k in model_dict}
#     # 2. overwrite entries in the existing state dict
#     model_dict.update(full_dict) 
#     # 3. load the new state dict
#     mels_model.load_state_dict(full_dict)
    
#     # Forward
#     with torch.no_grad():
#         mels_model.eval()
#         batch_output_dict = mels_model(x_mels, None)
    
#     framewise_output = batch_output_dict['framewise_output'].data.cpu().numpy()[0]
#     """(time_steps, classes_num)"""
    
#     print('Sound event detection result (time_steps x classes_num): {}'.format(
#         framewise_output.shape))
    
#     sorted_indexes = np.argsort(np.max(framewise_output, axis=0))[::-1]
    
#     top_k = 10  # Show top results
#     top_result_mat = framewise_output[:, sorted_indexes[0 : top_k]]    
#     """(time_steps, top_k)"""
    
#     return(framewise_output, labels, top_result_mat, sorted_indexes)
    
    # Plot result    
    # stft = librosa.core.stft(y=waveform[0].data.cpu().numpy(), n_fft=window_size, 
    #     hop_length=hop_size, window='hann', center=True)
    # frames_num = stft.shape[-1]
    
    # # Plot result    
    # stft = librosa.core.stft(y=waveform[0].data.cpu().numpy(), n_fft=window_size, 
    #     hop_length=hop_size, window='hann', center=True)
    # frames_num = stft.shape[-1]
    
    # fig, axs = plt.subplots(2, 1, sharex=True, figsize=(10, 4))
    # axs[0].matshow(np.log(np.abs(stft)), origin='lower', aspect='auto', cmap='jet')
    # axs[0].set_ylabel('Frequency bins')
    # axs[0].set_title('Log spectrogram')
    # axs[1].matshow(top_result_mat.T, origin='upper', aspect='auto', cmap='jet', vmin=0, vmax=1)
    # axs[1].xaxis.set_ticks(np.arange(0, frames_num, frames_per_second))
    # axs[1].xaxis.set_ticklabels(np.arange(0, frames_num / frames_per_second))
    # axs[1].yaxis.set_ticks(np.arange(0, top_k))
    # axs[1].yaxis.set_ticklabels(np.array(labels)[sorted_indexes[0 : top_k]])
    # axs[1].yaxis.grid(color='k', linestyle='solid', linewidth=0.3, alpha=0.3)
    # axs[1].set_xlabel('Seconds')
    # axs[1].xaxis.set_ticks_position('bottom')
    
    # plt.tight_layout()
    # plt.savefig('test')

# def from_xlsx_to_list(workbook_path):
#     workbook = openpyxl.load_workbook(workbook_path)
#     worksheet = workbook.active
#     first_row = [] # The row where we stock the name of the column
#     for col in range(worksheet.max_column):
#         first_row.append( worksheet.cell(0+1,col+1).value )
#     # tronsform the workbook to a list of dictionnary
#     data =[]
#     for row in range(1, worksheet.max_row):
#         elm = {}
#         for col in range(worksheet.max_column):
#             elm[first_row[col]]=worksheet.cell(row+1,col+1).value
#         data.append(elm)
#     return(data)

